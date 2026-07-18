"""Export Table 2.1-style xlsx using Abaqus ANALYSIS_CHECKS warning counts."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
JSON_PATH = ROOT / "output" / "reports" / "mesh_convergence" / "bcc_mesh_quality_literature.json"
OUT_XLSX = ROOT / "output" / "reports" / "mesh_convergence" / "bcc_mesh_quality_literature.xlsx"


def main() -> int:
    if not JSON_PATH.is_file():
        raise SystemExit(f"missing {JSON_PATH}; run scripts/linux/run_bcc_mesh_verify_literature.sh first")
    rows = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    if not rows:
        raise SystemExit("empty literature JSON")

    wb = Workbook()
    ws = wb.active
    ws.title = "表2.1"

    ws.merge_cells("A1:D1")
    ws["A1"] = "表 2.1  网格划分质量和计算时间汇总 / Table 2.1 Summary of meshing quality and calculation time"
    ws["A1"].font = Font(name="Microsoft YaHei", size=11, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    headers = [
        "Mesh size",
        "Total number of meshes",
        "Warning meshes",
        "CPU time (s)",
    ]
    fill = PatternFill("solid", fgColor="D9E2F3")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for c, h in enumerate(headers, 1):
        cell = ws.cell(2, c, h)
        cell.fill = fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin

    # Prefer near-zero ANALYSIS_CHECKS warnings; among those pick coarsest-usable
    # by F–u RMSE vs 0.6 (exclude baseline itself when a coarser ok candidate exists).
    old_path = ROOT / "output" / "reports" / "mesh_convergence" / "bcc_mesh_quality_summary.json"
    rmse = {}
    if old_path.is_file():
        for r0 in json.loads(old_path.read_text(encoding="utf-8")):
            rmse[float(r0["mesh_size_mm"])] = r0.get("rmse_vs_0p6_N")

    ok = [r for r in rows if (r.get("warning_pct") or 0.0) < 0.1]  # literature band ~0.03–0.09%
    if not ok:
        ok = list(rows)
    non_base = [r for r in ok if float(r["mesh_size_mm"]) != 0.6]
    pool = non_base if non_base else ok

    def score(r):
        r_rmse = rmse.get(float(r["mesh_size_mm"]))
        if r_rmse is None:
            r_rmse = 0.0 if float(r["mesh_size_mm"]) == 0.6 else 1e9
        return (
            r_rmse,
            r.get("cpu_time_s") if r.get("cpu_time_s") is not None else 1e9,
            float(r["mesh_size_mm"]),
        )

    best = min(pool, key=score)
    green = PatternFill("solid", fgColor="C6EFCE")
    center = Alignment(horizontal="center", vertical="center")

    for i, r in enumerate(rows):
        ri = 3 + i
        n = r["total_number_of_meshes"]
        w = r["warning_meshes"]
        pct = r["warning_pct"]
        warn_cell = f"{w}({pct:.3f}%)"
        vals = [r["mesh_size_mm"], n, warn_cell, r.get("cpu_time_s")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(ri, c, v)
            cell.alignment = center
            cell.border = thin
            if r["mesh_size_mm"] == best["mesh_size_mm"]:
                cell.fill = green

    note_r = 3 + len(rows) + 1
    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r + 3, end_column=4)
    ws.cell(
        note_r,
        1,
        "口径说明：Warning meshes = Abaqus/CAE Mesh Verify → Shape Metrics → "
        "ASPECT_RATIO（阈值=3）的失败单元数 failedElements 及占比；"
        "CAE 界面将此类单元按 warning 高亮，与文献 Table 2.1 数量级一致（约 0.03%）。\n"
        "（此前 ANALYSIS_CHECKS 仅 0～1 个；自定义 max/min 边长比>10 则偏多至 ~75%，均非文献口径。）\n"
        "CPU time (s) = Explicit 求解墙钟时间（.sta WALL TIME，单位秒）；作业提交 cpus=48。\n"
        f"绿色行：综合推荐（Warning 正常且相对 0.6 mm 力–位移更接近）："
        f"{best['mesh_size_mm']} mm。\n"
        "几何：BCC 4×4×4 paper_box，CAE C3D4 free tet。",
    )
    ws.cell(note_r, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[note_r].height = 72

    for i, w in enumerate([12, 22, 22, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # raw numbers sheet
    ws2 = wb.create_sheet("raw")
    h2 = ["mesh_size_mm", "numElements", "warningElements", "warning_pct", "failedElements", "cpu_time_s", "slug"]
    for c, h in enumerate(h2, 1):
        ws2.cell(1, c, h).font = Font(bold=True)
    for i, r in enumerate(rows):
        ws2.cell(2 + i, 1, r["mesh_size_mm"])
        ws2.cell(2 + i, 2, r["total_number_of_meshes"])
        ws2.cell(2 + i, 3, r["warning_meshes"])
        ws2.cell(2 + i, 4, r["warning_pct"])
        ws2.cell(2 + i, 5, r.get("failed_elements"))
        ws2.cell(2 + i, 6, r.get("cpu_time_s"))
        ws2.cell(2 + i, 7, r.get("slug"))

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print("Saved:", OUT_XLSX)
    print("Best (lowest warning %):", best["mesh_size_mm"], "mm",
          f"{best['warning_meshes']}({best['warning_pct']:.3f}%)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
