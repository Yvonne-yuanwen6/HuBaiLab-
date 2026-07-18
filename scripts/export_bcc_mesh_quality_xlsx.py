"""Build Table 2.1-style BCC mesh quality / timing workbook for mesh-size choice."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
JSON_PATH = ROOT / "output" / "reports" / "mesh_convergence" / "bcc_mesh_quality_summary.json"
OUT_XLSX = ROOT / "output" / "reports" / "mesh_convergence" / "bcc_mesh_quality_summary.xlsx"


def main() -> int:
    rows = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    wb = Workbook()

    # ---- Sheet 1: Table 2.1 style ----
    ws = wb.active
    ws.title = "表2.1_网格与时间"

    title = "表 2.1  BCC 4×4×4 网格划分质量与计算时间汇总（CAE C3D4）"
    ws.merge_cells("A1:G1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Microsoft YaHei", size=12, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    headers = [
        "Mesh size (mm)",
        "Total number of meshes",
        "Warning meshes\n(aspect>10 in sample)",
        "Warning ratio",
        "Wall time (s)",
        "Wall time (hh:mm:ss)",
        "CPU time est. (s)\n(= wall × 48)",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(2, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[2].height = 36

    # Recommend 0.8: highlight row
    recommend_seed = 0.8
    # score: minimize rmse vs 0.6 and wall time, avoid 1.2 peak jump
    best = min(
        (r for r in rows if r["mesh_size_mm"] != 0.6),
        key=lambda r: (
            (r.get("rmse_vs_0p6_N") or 1e9),
            r.get("wall_time_s") or 1e9,
            abs((r.get("peak_force_N") or 0) - rows[0].get("peak_force_N", 0)),
        ),
    )
    recommend_seed = best["mesh_size_mm"]

    center = Alignment(horizontal="center", vertical="center")
    highlight = PatternFill("solid", fgColor="C6EFCE")
    for i, r in enumerate(rows):
        row_i = 3 + i
        warn = r.get("warning_meshes")
        sample_n = r.get("warning_sample_n") or 0
        warn_str = f"{warn} ({r['warning_pct']:.3f}% of sample)" if warn is not None else "—"
        ratio_str = f"{r['warning_pct']:.3f}%" if r.get("warning_pct") is not None else "—"
        values = [
            r["mesh_size_mm"],
            r.get("nelem_c3d4") or r.get("nelem_total"),
            warn_str,
            ratio_str,
            r.get("wall_time_s"),
            r.get("wall_time"),
            r.get("cpu_time_est_s"),
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row_i, col, v)
            cell.alignment = center
            cell.border = thin
            cell.font = Font(name="Calibri", size=10)
            if r["mesh_size_mm"] == recommend_seed:
                cell.fill = highlight

    ws.append([])
    note_row = 3 + len(rows) + 1
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 4, end_column=7)
    notes = (
        "说明：\n"
        "1) Total number of meshes = 压缩 INP 中 C3D4 单元数（含压板相关固体单元极少）。\n"
        "2) Warning meshes 来自 CAE pilot 对 aspect ratio 的抽样：sample≤120000，统计 aspect>10 的单元数及占抽样比例；"
        "与文献 Table 2.1 中 Abaqus “warning meshes”定义不同，仅作相对质量对比。\n"
        "3) Wall time 取自 .sta 最后一行 WALL TIME；CPU time est. = wall × 48（提交时 cpus=48）。\n"
        f"4) 绿色高亮为综合推荐网格尺寸：{recommend_seed} mm"
        "（相对 0.6 mm 的力–位移 RMSE 较小，峰力更稳，耗时不显著增加）。\n"
        "5) 1.2 mm 档使用 quality=fast 且杆径加密更少，单元数反而升高、峰力偏离 0.6 mm 最大，不推荐。"
    )
    ws.cell(note_row, 1, notes)
    ws.cell(note_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[note_row].height = 90

    widths = [14, 22, 28, 14, 14, 18, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Sheet 2: decision metrics ----
    ws2 = wb.create_sheet("选型辅助指标")
    h2 = [
        "Mesh size (mm)",
        "Elements (C3D4)",
        "Aspect p95",
        "Peak force (N)",
        "RMSE vs 0.6 mm (N)",
        "Wall time (s)",
        "Relative wall vs 0.6",
        "Recommended?",
    ]
    for col, h in enumerate(h2, 1):
        cell = ws2.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws2.row_dimensions[1].height = 32

    wall0 = rows[0].get("wall_time_s") or 1
    for i, r in enumerate(rows):
        row_i = 2 + i
        rel = (r.get("wall_time_s") or 0) / wall0
        rec = "YES" if r["mesh_size_mm"] == recommend_seed else ("baseline" if r["mesh_size_mm"] == 0.6 else "")
        vals = [
            r["mesh_size_mm"],
            r.get("nelem_c3d4"),
            r.get("aspect_p95"),
            None if r.get("peak_force_N") is None else round(r["peak_force_N"], 1),
            None if r.get("rmse_vs_0p6_N") is None else round(r["rmse_vs_0p6_N"], 2),
            r.get("wall_time_s"),
            round(rel, 3),
            rec,
        ]
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(row_i, col, v)
            cell.alignment = center
            cell.border = thin
            if r["mesh_size_mm"] == recommend_seed:
                cell.fill = highlight

    for i, w in enumerate([14, 16, 12, 14, 18, 14, 16, 14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ---- Sheet 3: recommendation ----
    ws3 = wb.create_sheet("推荐结论")
    ws3["A1"] = "最适合网格大小（基于本次 BCC 扫描）"
    ws3["A1"].font = Font(name="Microsoft YaHei", size=13, bold=True)
    ws3.merge_cells("A1:B1")
    ws3["A3"] = "推荐 mesh size"
    ws3["B3"] = f"{recommend_seed} mm"
    ws3["B3"].font = Font(name="Calibri", size=16, bold=True, color="006100")
    ws3["A4"] = "理由"
    reasons = [
        f"相对最细网 0.6 mm：RMSE={best.get('rmse_vs_0p6_N'):.2f} N（四档中最小非零档之一）",
        f"墙钟时间 {best.get('wall_time')}（约 {best.get('wall_time_s')} s），较 0.6 mm 略快或接近",
        f"峰值力 {best.get('peak_force_N'):.1f} N，与 0.6 mm 的 {rows[0].get('peak_force_N'):.1f} N 更接近",
        "力–位移全曲线与 0.6–1.0 mm 整体重合（见 bcc_quasi_static_mesh_validation.png）",
        "1.2 mm 峰力偏高、与细网偏离最大，不建议作为正式工况",
    ]
    ws3["B4"] = "\n".join(f"• {x}" for x in reasons)
    ws3["B4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws3.row_dimensions[4].height = 110
    ws3["A6"] = "若更强调精度"
    ws3["B6"] = "选用 0.6 mm（基线，最细，耗时最长）"
    ws3["A7"] = "若更强调速度且可接受更大误差"
    ws3["B7"] = "可试用 1.0 mm（墙钟最短），但密实化段峰力已开始偏离"
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 78

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Saved: {OUT_XLSX}")
    print(f"Recommended mesh size: {recommend_seed} mm")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
