#!/usr/bin/env python3
"""Build xlsx checklist of ambiguous thesis FE settings for author inquiry."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "论文复现_模糊设置清单_问作者.xlsx"

HEADERS = ["编号", "软件", "论文章节", "论文已明确", "仍模糊/未闭合", "建议向作者确认的问题"]

SHARED = [
    (
        "G1",
        "共用",
        "§2.1",
        "L=20 mm，杆径 d=2 mm，Q∈{0,0.5,1,1.5}，4×4×4",
        "Fig.2.6 实体 CAD 精确构造：杆端是 RVE 平面截断还是球头/节点球？枢纽处是否允许几何相交？",
        "Fig.2.6 有限元几何是扫掠圆管 + L³ 盒切割、无节点球，是否正确？能否提供原始 STEP/Parasolid？",
    ),
    (
        "G2",
        "共用",
        "§2.1 / 式(2.1)",
        "SFBLS 正弦屈曲公式 f(s)=A_f·sin(2πQs)",
        "A_f 是否恒为 2 mm？屈曲参考方向如何取？Q=1 时 sin 项符号/相位是否与 Q=0.5 相同？",
        "SFBLS 八根杆的屈曲方向是按胞心→角点向外凸，还是固定 Z 轴投影？Q=1 是否有特殊取向？",
    ),
    (
        "G3",
        "共用",
        "§2.1",
        "4×4×4 块体",
        "全局坐标系：压缩方向、堆叠方向、隔振激励方向（Y 还是 Z）在 CAD/FE 中如何统一？",
        "压缩 FE 与隔振 FE/试验是否共用同一坐标定义？激励 Y 轴与点阵堆叠轴的关系？",
    ),
    (
        "G4",
        "共用",
        "§2.3.2",
        "拉伸试验得 E=25 MPa，屈服 4.69 MPa，ρ=1135 kg/m³；Fig.2.5 有完整 σ–ε 曲线",
        "压缩 FE（§2.4.1）与隔振 FE（§2.4.3）是否共用同一本构输入？Fig.2.5 曲线是否分别用于 Abaqus 与 COMSOL？",
        "压缩仿真用的是 E/ν/ρ 线性参数，还是 Fig.2.5 超弹性/Marlow？隔振是否必须用 Marlow？",
    ),
    (
        "G5",
        "共用",
        "§2.3.2 / §2.4.1",
        "拉伸有屈服点 4.69 MPa",
        "压缩 FE 是否含塑性/损伤？屈服应力是否进入 Abaqus 材料卡？",
        "§2.4.1 是否仅超弹性/线弹性，还是弹塑性？若弹塑性，硬化律与压缩曲线来源？",
    ),
    (
        "G6",
        "共用",
        "§2.3.2",
        "Fig.2.5 为单轴拉伸",
        "Marlow/Ogden 等拟合时是否只用单轴数据？是否考虑压缩与单轴差异？",
        "超弹性标定是否仅 Fig.2.5 单轴点？压缩大变形是否另做双轴/平面应变数据？",
    ),
]

ABAQUS = [
    (
        "A1",
        "Abaqus",
        "§2.4.1",
        "Abaqus/Explicit；刚体顶/底板；底固定、顶仅竖向；μ=0.1 + 法向硬接触；C3D4；全局 0.6 mm；示例步长 0.64 s；准静态 KE/IE < 5%",
        "0.64 s 对应多大位移/应变？是完整算例还是示意片段？",
        "Fig.2.6 中 0.64 s 步长对应的压板行程和工程应变是多少？",
    ),
    (
        "A2",
        "Abaqus",
        "§2.4.2 vs §2.4.1",
        "试验十字头 5 mm/min；致密化停止（无固定应变）",
        "FE 加载速率是否也是 5 mm/min？FE 终止应变/行程如何取？",
        "§2.4.1 有限元是否采用 5 mm/min？仿真压到多少应变停止，是否按致密化判据自动停？",
    ),
    (
        "A3",
        "Abaqus",
        "§2.4.1",
        "C3D4、0.6 mm",
        "局部加密规则（杆径方向单元数、枢纽/接触区是否细化）？是否用过 C3D10/C3D8R？",
        "除全局 0.6 mm 外，是否有按杆径 d/单元数或曲率加密的规则？最终收敛网格单元类型与尺寸？",
    ),
    (
        "A4",
        "Abaqus",
        "§2.4.1",
        "刚体板 + 硬接触",
        "刚体板尺寸（相对 4×4×4 footprint 的余量、厚度）？板–点阵初始间隙/过盈如何处理？",
        "上下刚体板 XY 尺寸和厚度？初始是否零间隙、小过盈还是干涉装配？",
    ),
    (
        "A5",
        "Abaqus",
        "§2.4.1",
        "板–点阵接触",
        "点阵杆件之间是否启用自接触（大变形后防穿模）？若启用，摩擦与过盈策略？",
        "压缩 FE 是否包含 ALL EXTERIOR 自接触？枢纽初始过盈如何处理？",
    ),
    (
        "A6",
        "Abaqus",
        "§2.4.1",
        "Explicit + KE/IE 准则",
        "时间增量 Δt、质量缩放、幅值 ramp/hold 的具体取值？自适应还是固定步长？",
        "Explicit 的 Δt、质量缩放因子、加载幅值前段 hold 比例各是多少？",
    ),
    (
        "A7",
        "Abaqus",
        "§2.4.1",
        "TPU：ρ=1135 kg/m³，E=25 MPa，ν=0.47",
        "本构模型类型（Neo-Hooke / Marlow / Ogden / 多项式 / 线弹性）及完整参数表",
        "§2.4.1 材料卡完整截图或 INP 片段？是否用 Fig.2.5 试验数据输入？",
    ),
    (
        "A8",
        "Abaqus",
        "§2.4.1 / §2.4.2",
        "试验记录载荷 F、行程 S",
        "后处理定义：工程应力 σ 的分母是 nx·L·ny·L 还是实际承载投影面积？ε 如何定义？",
        "Fig.3.3 应力–应变曲线的 σ、ε 定义与 ODB 提取位置（板反力/位移还是单胞尺度）？",
    ),
    (
        "A9",
        "Abaqus",
        "§2.4.1",
        "KE/IE < 5%",
        "该准则需全程满足还是稳态段/峰值前即可？是否报告 KE/IE 曲线？",
        "KE/IE<5% 是全程约束还是仅平台段？若超限如何调整 Δt/质量缩放？",
    ),
    (
        "A10",
        "Abaqus",
        "Fig.3.3 / 3.11 / 3.13",
        "多 Q、多结构对比曲线",
        "各子图对应的阵列规模（3×3×3 还是 4×4×4）？BCC 与 SFBLS 是否同一网格/本构设置？",
        "Fig.3.3 每个算例的 nx×ny×nz、材料模型、网格是否完全一致？能否给算例清单？",
    ),
    (
        "A11",
        "Abaqus",
        "Fig.3.13",
        "能量吸收对比",
        "能量吸收积分区间（0→ε_max？是否含卸载？）与单位",
        "能量吸收是对 σ–ε 曲线哪一段积分？ε_max 取多少？",
    ),
]

COMSOL = [
    (
        "C1",
        "COMSOL",
        "§2.4.3",
        "COMSOL 5.6；固体力学；振动台 AISI 4340；输出端铝合金；点阵 TPU（Fig.2.5→Marlow）；ρ=1135 kg/m³",
        "COMSOL 内置材料库的具体条目名（4340 牌号、铝合金牌号）是否与试验件一致？",
        "振动台钢、铝片在 COMSOL 中选的是哪条材料库记录？是否有自定义参数？",
    ),
    (
        "C2",
        "COMSOL",
        "Fig.2.8",
        "400×400×400 mm 钢块；点阵 + 薄铝片",
        "铝片平面尺寸（相对点阵 footprint 外扩多少 mm）？0.5 mm 厚度是否即试验值？",
        "Fig.2.8 铝片边长是 80 mm、100 mm 还是其他？厚度 0.5 mm 是否实测？",
    ),
    (
        "C3",
        "COMSOL",
        "Table 3.3",
        "模态试验含 300 g 载荷",
        "300 g 在 FE 中建模为集中质量、均布面密度还是实体块？尺寸与位置？",
        "300 g 在 COMSOL 里是 Added Mass 还是几何实体？质心在哪？",
    ),
    (
        "C4",
        "COMSOL",
        "§2.4.3",
        "物理场控制网格：点阵细化、台/板常规",
        "「细化/常规」对应的 hmax 或 COMSOL hauto 数值？台顶接触区 8 mm 渐变是否为原文设定？",
        "点阵、振动台、铝片各自 hmax 或 Size 预设等级？接触区是否单独加密？",
    ),
    (
        "C5",
        "COMSOL",
        "§2.4.3",
        "固体力学离散",
        "单元阶次（线性/二次）与几何非线性是否开启？大变形是否启用？",
        "Solid Mechanics 用线性还是二次单元？Geometric nonlinearity 开/关？",
    ),
    (
        "C6",
        "COMSOL",
        "§2.4.3 / Fig.2.8",
        "振动台 + 点阵 + 铝片装配",
        "界面连接：共节点、Identity pair、Tie、Contact 还是 Form Assembly 粘结？摩擦有无？",
        "点阵–钢台–铝片界面是 bonded contact、identity pair 还是共享拓扑？",
    ),
    (
        "C7",
        "COMSOL",
        "§2.4.3",
        "振动台底面固定",
        "本征与谐响应是否同一套约束？铝片顶面是否额外约束？",
        "除台底 Fixed 外，是否还有其他边界（对称、弹性支撑）？",
    ),
    (
        "C8",
        "COMSOL",
        "§2.4.3",
        "指定加速度 0.98 m/s² 正弦；Y 轴",
        "激励施加在哪一面/哪一节点集（台顶面还是台底）？频域里是加速度幅值还是位移幅值？",
        "0.98 m/s² 是 Prescribed Acceleration 在台顶的幅值吗？Y 轴与 CAD 坐标如何对应？",
    ),
    (
        "C9",
        "COMSOL",
        "§2.4.3",
        "频域谐响应",
        "扫频范围与步长（Table 3.3 / Fig.3.22 用的 f_min、f_max、Δf）？",
        "谐响应扫频是 10–2000 Hz 还是其他？步长 1/5/10 Hz？",
    ),
    (
        "C10",
        "COMSOL",
        "§2.4.3 / Table 3.3",
        "报告 f_n、传递率",
        "阻尼模型（结构阻尼 η、Q 因子、Rayleigh、损耗因子）及取值？无阻尼本征 vs 有阻尼谐响应如何衔接？",
        "Table 3.3 传递率计算是否含阻尼？η 或 Q 取多少？本征频率是否无阻尼提取？",
    ),
    (
        "C11",
        "COMSOL",
        "式(3.20) / Table 3.3",
        "T = A_out / A_in",
        "A_out、A_in 的测点位置（台顶中心？铝片顶中心？平均面？）与加速度方向（仅 Y 还是合成）？",
        "传递率探针具体坐标/选面？A 是法向分量还是合成幅值？",
    ),
    (
        "C12",
        "COMSOL",
        "Table 3.3",
        "一阶 f_n（如 BCC 14.8 Hz）",
        "本征求解器设置：求多少阶模态？shift 频率？是否排除刚体/约束模态？",
        "Eigenfrequency 求几阶？shift-invert 的 shift 取多少 Hz？",
    ),
    (
        "C13",
        "COMSOL",
        "§2.4.3",
        "Marlow 来自 Fig.2.5",
        "隔振 FE 中 Marlow 是否仅用拉伸曲线？是否设泊松比/体积模量约束？",
        "COMSOL Hyperelastic Marlow 输入是 Fig.2.5 全部点还是截断应变？ν 如何处理？",
    ),
    (
        "C14",
        "COMSOL",
        "Fig.3.22",
        "传递率曲线",
        "曲线是仿真还是试验？若为仿真，是否与 Table 3.3 同一模型（含 300 g）？",
        "Fig.3.22 每条曲线对应 Q=0/0.5/1/1.5 时是否都含 Fig.2.8 全装配？",
    ),
]

CROSS = [
    (
        "X1",
        "后处理",
        "§2.2 vs §2.4",
        "理论用单胞面积/高度（式 2.11–2.12）",
        "Fig.3.3 宏观 σ–ε 用块体名义尺寸还是单胞？",
        "图表应力应变是 RVE 名义值还是试验 MTS 直接值？",
    ),
    (
        "X2",
        "共用",
        "两软件",
        "Abaqus 做压缩，COMSOL 做隔振",
        "同一 Q、同一 4×4×4 几何是否同一 CAD 源文件？",
        "两章 FE 是否共用同一阵列 STEP？能否提供各 Q 的官方几何？",
    ),
    (
        "X3",
        "对照",
        "Table 3.3 vs Fig.3.3",
        "不同物理问题",
        "BCC Q=0 在 Table 3.3 的 f_n 与 Fig.3.3 压缩曲线是否同一批次样件/参数？",
        "隔振与压缩算例的材料批次、几何 Q 定义是否完全一致？",
    ),
]

PRIORITY = [
    (
        "P1",
        "共用",
        "§2.1 / Fig.2.6",
        "—",
        "Fig.2.6 CAD 定义",
        "无节点球 + L³ 平面切割是否正确？能否提供 STEP？",
    ),
    (
        "P2",
        "Abaqus",
        "§2.4.1",
        "—",
        "材料本构",
        "Abaqus 完整材料模型与参数（是否用 Fig.2.5 数据）？",
    ),
    (
        "P3",
        "Abaqus",
        "§2.4.1 / §2.4.2",
        "—",
        "加载条件",
        "FE 速率是否 5 mm/min？终止应变/致密化判据？",
    ),
    (
        "P4",
        "Abaqus",
        "§2.4.1 / Fig.2.6",
        "—",
        "0.64 s 算例",
        "对应位移/应变及 KE/IE 是否全程 <5%？",
    ),
    (
        "P5",
        "Abaqus",
        "§2.4.1",
        "—",
        "自接触",
        "杆件自接触是否启用？初始过盈如何处理？",
    ),
    (
        "P6",
        "Abaqus",
        "§2.4.1",
        "—",
        "Explicit 数值",
        "Δt、质量缩放、幅值 ramp 具体值？",
    ),
    (
        "P7",
        "Abaqus",
        "Fig.3.3",
        "—",
        "σ–ε 定义",
        "Fig.3.3 的 σ、ε 计算公式与提取位置？",
    ),
    (
        "P8",
        "COMSOL",
        "§2.4.3",
        "—",
        "坐标系",
        "激励 Y 轴与点阵堆叠/压缩 Z 轴的关系？",
    ),
    (
        "P9",
        "COMSOL",
        "§2.4.3",
        "—",
        "阻尼与扫频",
        "谐响应阻尼取值、扫频范围与步长？",
    ),
    (
        "P10",
        "COMSOL",
        "式(3.20)",
        "—",
        "传递率探针",
        "A_in、A_out 的具体测面/测点与方向？",
    ),
]

META = [
    ("文献", "Hu & Bai (2024) 博士论文"),
    ("用途", "复现前向原作者确认的模糊/未闭合 FE 设置（不含本仓库工程推断）"),
    ("生成", "scripts/build_thesis_ambiguity_xlsx.py"),
    ("说明", "「论文已明确」摘自论文章节；「仍模糊」为原文未写清或存在多种解读之处"),
]


def _style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF", size=11)
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_sheet(ws, title: str, rows: list[tuple[str, ...]]) -> None:
    ws.title = title
    ws.append(HEADERS)
    _style_header(ws)
    for row in rows:
        ws.append(list(row))
    widths = [8, 10, 14, 36, 40, 44]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(HEADERS) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(
                vertical="top", wrap_text=True
            )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"


def main() -> None:
    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "说明"
    ws_meta.append(["字段", "内容"])
    _style_header(ws_meta)
    for k, v in META:
        ws_meta.append([k, v])
    ws_meta.column_dimensions["A"].width = 12
    ws_meta.column_dimensions["B"].width = 72
    for r in range(2, ws_meta.max_row + 1):
        ws_meta.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    sheets = [
        ("共用_几何材料", SHARED),
        ("Abaqus_压缩", ABAQUS),
        ("COMSOL_隔振", COMSOL),
        ("跨软件_后处理", CROSS),
        ("优先问作者10条", PRIORITY),
    ]
    for name, data in sheets:
        _write_sheet(wb.create_sheet(name), name, data)

    # 汇总 sheet
    ws_all = wb.create_sheet("全部汇总", 1)
    ws_all.append(HEADERS + ["分类"])
    _style_header(ws_all)
    for cat, data in [
        ("共用_几何材料", SHARED),
        ("Abaqus_压缩", ABAQUS),
        ("COMSOL_隔振", COMSOL),
        ("跨软件_后处理", CROSS),
    ]:
        for row in data:
            ws_all.append(list(row) + [cat])
    widths = [8, 10, 14, 36, 40, 44, 14]
    for i, w in enumerate(widths, start=1):
        ws_all.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, ws_all.max_row + 1):
        for c in range(1, len(HEADERS) + 2):
            ws_all.cell(row=r, column=c).alignment = Alignment(
                vertical="top", wrap_text=True
            )
    ws_all.freeze_panes = "A2"
    ws_all.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS)+1)}{ws_all.max_row}"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT} ({ws_all.max_row - 1} items + 10 priority)")


if __name__ == "__main__":
    main()
