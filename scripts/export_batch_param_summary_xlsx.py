#!/usr/bin/env python3
"""Export batch lattice structural-parameter summary xlsx (with fixed image slots)."""

from __future__ import annotations

import json
import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_points

ROOT = Path(__file__).resolve().parents[1]
INDEX_PATH = ROOT / "output" / "cad" / "批量构型" / "_batch_index.json"
OUT_XLSX = ROOT / "output" / "reports" / "批量构型结构参数汇总.xlsx"

IMG_PX = 160
L_MM = 20.0
ARRAY = "4×4×4"

PHASE_LABEL = {
    "A_circle": "A·圆杆",
    "A_ellipse_k2": "A·椭圆κ=2",
    "A_ellipse_k1p5": "A·椭圆κ=1.5",
    "B_Af": "B·Af扫参",
    "B_area": "B·截面积扫参",
}


def _width_from_pixels(px: int, mdw: float = 7.0) -> float:
    return max(1.0, (float(px) - 5.0) / mdw)


def _profile(deq_mm: float, k: float) -> tuple[str, float, float]:
    deq = float(deq_mm)
    kappa = float(k)
    if abs(kappa - 1.0) < 1e-9:
        return "圆", deq, deq
    d_major = deq * math.sqrt(kappa)
    d_minor = deq / math.sqrt(kappa)
    return "椭圆(ellmin)", d_major, d_minor


def _fmt(v: float) -> float | int:
    if abs(v - round(v)) < 1e-9:
        return int(round(v))
    return round(v, 4)


def main() -> int:
    index = json.loads(INDEX_PATH.read_text(encoding="utf-8"))
    order = list(index.get("generation_order") or [])
    cases = dict(index.get("cases") or {})
    if not order:
        order = sorted(cases.keys())

    img_col_width = _width_from_pixels(IMG_PX)
    img_row_height = float(pixels_to_points(IMG_PX))

    wb = Workbook()
    ws = wb.active
    ws.title = "结构参数"

    title = (
        f"批量构型 · 结构参数汇总（paper_box · L=20 mm · 4×4×4）"
        f"　｜　构型图：右键单元格 → 粘贴选项 →「在单元格中粘贴图片」（勿用 Ctrl+V）"
    )
    headers = [
        "序号",
        "case_id",
        "组",
        "Af (mm)",
        "Q",
        "deq (mm)",
        "κ (k)",
        "截面",
        "长径 (mm)",
        "短径 (mm)",
        "L (mm)",
        "阵列",
        "构型图",
    ]
    n_cols = len(headers)
    img_col = n_cols

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c0 = ws.cell(1, 1, title)
    c0.font = Font(name="Microsoft YaHei", size=11, bold=True)
    c0.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Microsoft YaHei", color="FFFFFF", bold=True, size=10)
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    img_fill = PatternFill("solid", fgColor="F2F2F2")
    zebra = PatternFill("solid", fgColor="F7FBFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(2, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[2].height = 32

    for i, case_id in enumerate(order, 1):
        meta = cases.get(case_id) or {}
        af = float(meta["Af"])
        q = float(meta["Q"])
        deq = float(meta["deq_mm"])
        k = float(meta["k"])
        phase = str(meta.get("phase") or "")
        section, d_maj, d_min = _profile(deq, k)
        row = 2 + i
        values = [
            i,
            case_id,
            PHASE_LABEL.get(phase, phase),
            _fmt(af),
            _fmt(q),
            _fmt(deq),
            _fmt(k),
            section,
            _fmt(d_maj),
            _fmt(d_min),
            _fmt(L_MM),
            ARRAY,
            "",
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row, col, v)
            cell.alignment = center
            cell.border = thin
            cell.font = Font(name="Calibri" if col != 2 else "Consolas", size=10)
            if col == img_col:
                cell.fill = img_fill
            elif i % 2 == 0:
                cell.fill = zebra
        ws.row_dimensions[row].height = img_row_height

    widths = [6, 22, 14, 10, 8, 10, 8, 14, 10, 10, 8, 8, img_col_width]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    note_row = 3 + len(order)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 4, end_column=n_cols)
    notes = (
        "粘贴构型图（嵌入单元格，自动适应格子大小）：\n"
        "1) 先复制图片（截图/文件均可），再选中「构型图」列对应单元格。\n"
        "2) 右键该单元格 →「粘贴选项」里点「在单元格中粘贴图片」"
        "（开始 → 粘贴 ▼ 里也有同名项）。不要用 Ctrl+V（会浮在格子上并超出）。\n"
        "3) 若已用 Ctrl+V 贴成浮动图：选中图片 → 右键「置于单元格内」，"
        "或点图片旁出现的「置于单元格内」按钮。\n"
        "4) 从文件插入：插入 → 图片 →「置于单元格内」→ 此设备。\n"
        f"5) 参数来自 _batch_index.json；椭圆等面积 长径=deq√κ、短径=deq/√κ；"
        f"图槽约 {IMG_PX}×{IMG_PX} px。"
    )
    note_cell = ws.cell(note_row, 1, notes)
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    note_cell.font = Font(name="Microsoft YaHei", size=9, color="555555")
    ws.row_dimensions[note_row].height = 90

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}{2 + len(order)}"
    ws.print_title_rows = "1:2"

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX.relative_to(ROOT)} ({len(order)} cases)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
